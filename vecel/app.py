from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, Search
from datetime import datetime, timedelta
from functools import wraps
from flask_wtf.csrf import CSRFProtect
import os
import requests
import json
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import urllib3
from io import BytesIO
from openpyxl import Workbook
from forms import LoginForm, RegistrationForm
import pandas as pd
import urllib.parse
from dotenv import load_dotenv

# 환경 변수 로드
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'default-dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///site.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db.init_app(app)
csrf = CSRFProtect(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = '로그인이 필요한 페이지입니다.'

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Constants
TRADE_TYPE_MAPPING = {
    "전체": "",
    "매매": "A1",
    "전세": "B1",
    "월세": "B2"
}

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def check_subscription():
    if current_user.is_admin or current_user.is_premium:
        return True
    if current_user.subscription_expiry and datetime.utcnow() < current_user.subscription_expiry:
        return True
    return False

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, 
                   email=form.email.data, 
                   password=form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('회원가입이 완료되었습니다. 로그인해주세요.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html', form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember.data)
            user.last_login = datetime.utcnow()
            db.session.commit()
            flash('로그인되었습니다.', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('로그인 실패. 이메일과 비밀번호를 확인해주세요.', 'danger')
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    if not check_subscription():
        flash('체험 기간이 만료되었습니다. 관리자에게 문의해주세요.')
        return redirect(url_for('index'))
    return render_template('dashboard.html')

@app.route('/search', methods=['POST'])
@login_required
def search():
    """아파트 단지 검색 API"""
    if not current_user.is_active:
        return jsonify({'error': '계정이 비활성화되었습니다.'}), 403
        
    if not check_subscription():
        return jsonify({'error': '구독이 만료되었습니다.'}), 403
        
    try:
        # JSON 요청과 form 데이터 모두 처리
        if request.is_json:
            data = request.get_json()
            keyword = data.get('keyword', '').strip()
        else:
            keyword = request.form.get('keyword', '').strip()
        
        if not keyword:
            return jsonify({'error': '검색어를 입력해주세요.'}), 400
            
        print(f"[DEBUG] Searching for keyword: {keyword}")
        
        # 단지 검색
        complexes = get_complexes_by_region(keyword)
        if not complexes:
            print(f"[DEBUG] No complexes found for keyword: {keyword}")
            return jsonify({'error': '검색 결과가 없습니다.'}), 404
            
        # 응답 데이터 가공
        result = [{
            'complexNo': complex.get('complexNo'),
            'complexName': complex.get('complexName'),
            'address': f"{complex.get('address', '')} {complex.get('detailAddress', '')}".strip(),
            'totalHouseholdCount': complex.get('totalHouseholdCount'),
            'completionYearMonth': complex.get('completionYearMonth')
        } for complex in complexes if complex.get('complexNo')]
        
        if not result:
            print(f"[DEBUG] No valid complexes after processing")
            return jsonify({'error': '검색 결과가 없습니다.'}), 404
            
        print(f"[DEBUG] Found {len(result)} complexes")
        return jsonify({'complexes': result})
        
    except requests.exceptions.RequestException as e:
        print(f"[ERROR] Network error: {str(e)}")
        return jsonify({'error': '네트워크 오류가 발생했습니다.'}), 500
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error: {str(e)}")
        return jsonify({'error': '서버 응답을 처리할 수 없습니다.'}), 500
    except Exception as e:
        print(f"[ERROR] Unexpected error: {str(e)}")
        return jsonify({'error': f'검색 중 오류가 발생했습니다: {str(e)}'}), 500

def get_complexes_by_region(keyword):
    """네이버 부동산 API를 통해 아파트 단지 검색"""
    try:
        cookies = {
            'NNB': 'FGYNFS4Y6M6WO',
            'NFS': '2',
            'ASID': 'afd10077000001934e8033f50000004e',
            'ba.uuid': 'a5e52e8f-1775-4eea-9b42-30223205f9df',
            'tooltipDisplayed': 'true',
            'nstore_session': 'zmRE1M3UHwL1GmMzBg0gfcKH',
            '_fwb': '242x1Ggncj6Dnv0G6JF6g8h.1738045585397',
            'landHomeFlashUseYn': 'N',
            'REALESTATE': 'Thu Apr 03 2025 20:14:11 GMT+0900 (Korean Standard Time)',
            'NACT': '1',
        }
        headers = {
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3MzgwNDcxNjMsImV4cCI6MTczODA1Nzk2M30.Heq-J33LY9pJDnYOqmRhTTrSPqCpChtWxka_XUphnd4',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }
        params = {'keyword': keyword, 'page': '1'}
        url = 'https://new.land.naver.com/api/search'
        print(f"[DEBUG] 지역 검색 API 요청: {url} with params={params}")
        
        session = requests.Session()
        session.headers.update(headers)
        session.cookies.update(cookies)
        
        # 세션 초기화
        init_url = 'https://new.land.naver.com/complexes'
        init_response = session.get(init_url, verify=False, timeout=10)
        print(f"[DEBUG] Initial request status: {init_response.status_code}")
        if init_response.status_code != 200:
            print(f"[ERROR] Initial request failed: {init_response.text[:500]}")
            return []
            
        response = session.get(url, params=params, verify=False, timeout=10)
        print(f"[DEBUG] API response status: {response.status_code}")
        if response.status_code != 200:
            print(f"[ERROR] API request failed: {response.text[:500]}")
            return []
            
        # JSON 파싱 전 응답 내용 일부 로그
        response_text = response.text[:500]
        print(f"[DEBUG] API response preview: {response_text}")
        
        data = response.json()
        complexes = data.get('complexes', [])
        if not complexes:
            print(f"[DEBUG] No complexes in response")
            return []
            
        complexes.sort(key=lambda x: x['complexName'])
        print(f"[DEBUG] Found {len(complexes)} complexes")
        return complexes
    except requests.exceptions.RequestException as e:
        print(f"[ERROR] Request error: {str(e)}")
        return []
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error: {str(e)}, response: {response.text[:500] if 'response' in locals() else 'N/A'}")
        return []
    except Exception as e:
        print(f"[ERROR] API error: {str(e)}")
        return []

def get_real_estate_data(complex_no, trade_type, page=1):
    """매물 데이터 조회"""
    try:
        cookies = {
            'NNB': 'FGYNFS4Y6M6WO',
            'NFS': '2',
            'ASID': 'afd10077000001934e8033f50000004e',
            'ba.uuid': 'a5e52e8f-1775-4eea-9b42-30223205f9df',
            'tooltipDisplayed': 'true',
            'nstore_session': 'zmRE1M3UHwL1GmMzBg0gfcKH',
            'nstore_pagesession': 'iH4K+dqWcpYFllsM1U4-116496',
            'NAC': 'XfPpC4A0XeLCA',
            'page_uid': 'iHmGBsqVN8ossOXBRrlsssssswV-504443',
            'nhn.realestate.article.rlet_type_cd': 'A01',
            'nhn.realestate.article.trade_type_cd': '""',
            'nhn.realestate.article.ipaddress_city': '1100000000',
            '_fwb': '242x1Ggncj6Dnv0G6JF6g8h.1738045585397',
            'realestate.beta.lastclick.cortar': '1174010900',
            'landHomeFlashUseYn': 'N',
            'BUC': 'fwUJCqRUIsM47V0-Lcz1VazTR9EQgUrBIxM1P_x9Id4=',
            'REALESTATE': 'Tue Jan 28 2025 16:23:02 GMT+0900 (Korean Standard Time)',
            'NACT': '1',
        }
        headers = {
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3MzgwNDcxNjMsImV4cCI6MTczODA1Nzk2M30.Heq-J33LY9pJDnYOqmRhTTrSPqCpChtWxka_XUphnd4',
            'referer': f'https://new.land.naver.com/complexes/{complex_no}',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }
        url = f'https://new.land.naver.com/api/articles/complex/{complex_no}'
        params = {
            'realEstateType': 'APT:PRE:ABYG:JGC',
            'tradeType': trade_type,
            'tag': '::::::::',
            'rentPriceMin': '0',
            'rentPriceMax': '900000000',
            'priceMin': '0',
            'priceMax': '900000000',
            'areaMin': '0',
            'areaMax': '900000000',
            'showArticle': 'false',
            'sameAddressGroup': 'false',
            'priceType': 'RETAIL',
            'page': str(page),
            'complexNo': str(complex_no),
            'type': 'list',
            'order': 'rank'
        }
        print(f"[DEBUG] API 요청 시작: {url} with params={params}")
        session = requests.Session()
        session.headers.update(headers)
        session.cookies.update(cookies)
        
        # 세션 초기화
        init_url = f'https://new.land.naver.com/complexes/{complex_no}'
        init_response = session.get(init_url, verify=False, timeout=10)
        print(f"[DEBUG] Initial request status: {init_response.status_code}")
        if init_response.status_code != 200:
            print(f"[ERROR] Initial request failed: {init_response.text[:500]}")
            return None
            
        response = session.get(url, params=params, verify=False, timeout=10)
        print(f"[DEBUG] API response status: {response.status_code}")
        if response.status_code != 200:
            print(f"[ERROR] API request failed: {response.text[:500]}")
            return None
            
        data = response.json()
        print(f"[DEBUG] API 응답 성공: 상태 코드 {response.status_code}")
        return data
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error: {str(e)}, response: {response.text[:500] if 'response' in locals() else 'N/A'}")
        return None
    except Exception as e:
        print(f"[ERROR] Error: {str(e)}")
        return None

def process_data(data):
    """매물 데이터 처리"""
    if not data:
        print("[DEBUG] 처리할 데이터 없음")
        return []
    articles = data.get('articleList', [])
    if not articles:
        print("[DEBUG] articleList가 비어 있음")
    processed_data = []
    for article in articles:
        tags = ', '.join(article.get('tagList', [])) if article.get('tagList') else ''
        processed_article = {
            '순번': 'N/A',
            '아파트명': article.get('articleName'),
            '거래유형': article.get('tradeTypeName'),
            '층수': article.get('floorInfo'),
            '월세': article.get('rentPrc'),
            '거래가격': article.get('dealOrWarrantPrc'),
            '면적(m²)': article.get('area2'),
            '방향': article.get('direction'),
            '등록일': article.get('articleConfirmYmd'),
            '동': article.get('buildingName'),
            '중개사무소': article.get('realtorName'),
            '특징': tags
        }
        processed_data.append(processed_article)
    print(f"[DEBUG] 처리된 매물 수: {len(processed_data)}")
    return processed_data

def fetch_all_pages(complex_no, trade_type):
    """모든 페이지의 매물 데이터 수집"""
    all_data = []
    page = 1
    while True:
        print(f"[DEBUG] 데이터 수집 중... (페이지 {page})")
        response_data = get_real_estate_data(complex_no, trade_type, page)
        if not response_data or not response_data.get('articleList'):
            print(f"[DEBUG] 페이지 {page}에서 데이터 없음 또는 articleList 비어 있음")
            break
        processed_data = process_data(response_data)
        all_data.extend(processed_data)
        if not response_data.get('isMoreData', False):
            print(f"[DEBUG] 더 이상 데이터 없음")
            break
        page += 1
        time.sleep(1)
    all_data.sort(key=lambda x: x['등록일'] if x['등록일'] else '99999999')
    print(f"[DEBUG] 데이터 수집 완료. 총 {len(all_data)}개 매물 발견.")
    return all_data

@app.route('/fetch_data', methods=['POST'])
@login_required
def fetch_data():
    """매물 데이터 조회 API"""
    if not check_subscription():
        return jsonify({'error': '구독이 만료되었습니다.'}), 403
        
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '요청 데이터가 없습니다.'}), 400
            
        complex_no = data.get('complex_no')
        trade_type = data.get('trade_type', '')
        
        if not complex_no:
            return jsonify({'error': '단지 번호가 필요합니다.'}), 400
            
        # 거래 유형 코드 변환
        trade_type_code = TRADE_TYPE_MAPPING.get(trade_type, '')
        
        # 단지 정보 조회
        complex_info = get_complex_info(complex_no)
        if not complex_info:
            return jsonify({'error': '단지 정보를 찾을 수 없습니다.'}), 404
            
        # 매물 목록 조회
        article_list = get_article_list(complex_no, trade_type_code)
        if not article_list:
            return jsonify({'error': '매물 정보를 찾을 수 없습니다.'}), 404
            
        # 데이터 처리
        processed_data = []
        for article in article_list.get('articleList', []):
            price_info = get_price_info(article)
            
            # 층수 정보 처리
            floor_info = article.get('floorInfo', '')
            if floor_info:
                if '/' in floor_info:  # "3/15" 형태인 경우
                    current_floor = floor_info.split('/')[0]
                    floor_display = f"{current_floor}층"
                else:
                    floor_display = f"{floor_info}층"
            else:
                floor_display = '-'
            
            # 거래 유형이 선택되었고 "전체"가 아닌 경우, 해당 거래 유형만 필터링
            article_trade_type = article.get('tradeTypeName', '-')
            if trade_type and trade_type != '전체' and article_trade_type != trade_type:
                continue
            
            item = {
                '거래유형': article.get('tradeTypeName', '-'),
                '동': article.get('buildingName', '-'),
                '층수': floor_display,
                '전용면적': article.get('area2', '-'),
                '방향': article.get('direction', '-'),
                '거래가격': price_info['거래가격'],
                '월세': price_info['월세'],
                '중개사무소': article.get('realtorName', '-'),
                '등록일': article.get('articleConfirmYmd', '-'),
                '특징': ', '.join(article.get('tagList', [])) or '-',
                '특징설명': article.get('articleFeatureDesc', '')
            }
            processed_data.append(item)
            
        print(f"[DEBUG] 최종 가공된 데이터 수: {len(processed_data)}")
        return jsonify({'data': processed_data})
        
    except Exception as e:
        print(f"[ERROR] Error in fetch_data: {str(e)}")
        return jsonify({'error': f'데이터 조회 중 오류가 발생했습니다: {str(e)}'}), 500

@app.route('/download_excel', methods=['POST'])
def download_excel():
    try:
        data = request.get_json()
        complex_no = data.get('complex_no')
        trade_type = data.get('trade_type', '')
        
        if not complex_no:
            return jsonify({'error': '단지 번호가 필요합니다.'}), 400
            
        # 단지 정보 조회
        complex_info = get_complex_info(complex_no)
        if not complex_info:
            return jsonify({'error': '단지 정보를 찾을 수 없습니다.'}), 404
            
        # 매물 목록 조회
        article_list = get_article_list(complex_no, TRADE_TYPE_MAPPING.get(trade_type, ''))
        if not article_list or not article_list.get('articleList'):
            return jsonify({'error': '매물 정보를 찾을 수 없습니다.'}), 404
            
        # 거래 유형에 따라 매물 필터링
        filtered_articles = article_list['articleList']
        if trade_type and trade_type != '전체':
            filtered_articles = [article for article in filtered_articles 
                               if article.get('tradeTypeName') == trade_type]
            
            if not filtered_articles:
                return jsonify({'error': f'{trade_type} 유형의 매물이 없습니다.'}), 404
                
        print(f"[DEBUG] 필터링된 매물 수: {len(filtered_articles)}")
            
        # Excel 파일 생성
        wb = Workbook()
        ws = wb.active
        complex_name = complex_info.get('complexName', '매물정보')
        ws.title = complex_name
        
        # 아파트 이름 셀 병합 및 스타일 설정
        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = complex_name + (f' ({trade_type})' if trade_type and trade_type != '전체' else '')
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws.row_dimensions[1].height = 35  # 제목 행 높이 설정
        
        # 헤더 설정
        headers = ['순번', '거래유형', '동', '층수', '전용면적(㎡)', '방향', '거래가격', '월세', '중개사무소', '등록일', '특징']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30  # 헤더 행 높이 설정
        
        # 데이터 입력 및 최대 길이 계산
        max_widths = {i: len(header) for i, header in enumerate(headers, 1)}
        
        for idx, article in enumerate(filtered_articles, 1):
            price_info = get_price_info(article)
            floor_info = article.get('floorInfo', '')
            if floor_info:
                if '/' in floor_info:
                    current_floor = floor_info.split('/')[0]
                    floor_display = f"{current_floor}층"
                else:
                    floor_display = f"{floor_info}층"
            else:
                floor_display = '-'
            
            # 특징과 특징설명을 합쳐서 하나의 문자열로 만듦
            features = article.get('tagList', [])
            feature_desc = article.get('articleFeatureDesc', '')
            if feature_desc:
                features.append(feature_desc)
            features_text = ', '.join(filter(None, features)) or '-'
            
            row_data = [
                idx,  # 순번
                article.get('tradeTypeName', '-'),
                article.get('buildingName', '-'),
                floor_display,
                article.get('area2', '-'),
                article.get('direction', '-'),
                price_info.get('거래가격', '-'),
                price_info.get('월세', '-'),
                article.get('realtorName', '-'),
                article.get('articleConfirmYmd', '-'),
                features_text
            ]
            
            # 각 컬럼의 최대 길이 업데이트
            for col, value in enumerate(row_data, 1):
                str_value = str(value)
                cell = ws.cell(row=idx+2, column=col)
                cell.value = value
                cell.font = Font(size=10)
                
                # 최대 너비 업데이트
                if len(str_value) > max_widths.get(col, 0):
                    max_widths[col] = len(str_value)
                
                # 중앙 정렬 (특징 컬럼은 왼쪽 정렬)
                if col == 11:  # 특징 컬럼
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 설정
        for col, width in max_widths.items():
            col_letter = get_column_letter(col)
            # 특징 컬럼은 최대 50자 너비로 제한, 중개사무소 컬럼은 전체 텍스트 반영, 다른 컬럼은 최소 8자 이상으로 설정
            if col == 11:  # 특징 컬럼
                adjusted_width = min(width + 2, 50)
            elif col == 9:  # 중개사무소 컬럼
                adjusted_width = width + 2  # 중개사무소는 전체 텍스트 길이에 맞게 조정 (제한 없음)
            else:
                adjusted_width = max(width + 2, 8)
            
            ws.column_dimensions[col_letter].width = adjusted_width
            
        # 행 높이 설정 (모든 데이터 행의 높이를 30으로 고정)
        for row_idx in range(3, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30
            
        # 테두리 설정
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                
        # 자동 필터 설정
        ws.auto_filter.ref = f'A2:K{ws.max_row}'
            
        # 임시 파일로 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 현재 날짜와 시간을 파일명에 포함
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        suggested_filename = f'{complex_name}_{trade_type if trade_type and trade_type != "전체" else "전체"}_{current_time}.xlsx'
        
        response = send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=suggested_filename
        )
        
        # Content-Disposition 헤더 설정으로 다운로드 대화상자 표시
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(suggested_filename)}'
        
        return response
        
    except Exception as e:
        print(f"Excel download error: {str(e)}")
        return jsonify({'error': '엑셀 파일 생성 중 오류가 발생했습니다.'}), 500

def get_price_info(article):
    deal = article.get('dealOrWarrantPrc', '없음')
    rent = article.get('rentPrc', '없음')
    trade = article.get('tradeTypeName', '없음')

    print(f"[TEST] 거래유형: {trade} / 거래가: {deal} / 월세: {rent}")
    print(f"[TEST] 전체 article 데이터:")
    print(article)
    
    if trade == '월세':
        return {
            '거래가격': f"{deal}",
            '월세': f"{rent}"
        }
    else:
        return {
            '거래가격': f"{deal}",
            '월세': '-'
        }

def format_price(price):
    """가격을 포맷팅하는 함수"""
    try:
        if not price or price == '0':
            return '-'
            
        price = int(price)
        if price >= 10000:
            억 = price // 10000
            만 = price % 10000
            if 만 > 0:
                return f"{억}억 {만:,}만원"
            return f"{억}억원"
        else:
            return f"{price:,}만원"
    except:
        return '-'

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('관리자 권한이 필요합니다.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin')
@admin_required
def admin_dashboard():
    page = request.args.get('page', 1, type=int)
    users = User.query.paginate(page=page, per_page=10)
    total_users = User.query.count()
    active_users = User.query.filter_by(account_status='활성').count()
    trial_users = User.query.filter_by(is_premium=False).count()
    premium_users = User.query.filter_by(is_premium=True).count()
    
    return render_template('admin_dashboard.html',
                         users=users,
                         total_users=total_users,
                         active_users=active_users,
                         trial_users=trial_users,
                         premium_users=premium_users)

@app.route('/admin/toggle_subscription/<int:user_id>', methods=['POST'])
@admin_required
def toggle_subscription(user_id):
    user = User.query.get_or_404(user_id)
    if not user.is_admin:
        try:
            data = request.get_json()
            is_premium = data.get('is_premium', False)
            
            if is_premium != user.is_premium:
                user.is_premium = is_premium
                
                if is_premium:
                    # 유료 회원으로 설정 (약 100년)
                    user.subscription_expiry = datetime.utcnow() + timedelta(days=36500)  # 약 100년
                    flash(f'{user.username} 님을 유료 회원으로 변경했습니다.', 'success')
                else:
                    # 무료 체험으로 변경 (24시간)
                    user.subscription_expiry = datetime.utcnow() + timedelta(hours=24)
                    flash(f'{user.username} 님을 무료 체험 회원으로 변경했습니다.', 'success')
                
                db.session.commit()
                return jsonify({'success': True})
            else:
                return jsonify({'success': True, 'message': '변경사항이 없습니다.'})
                
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] 구독 유형 변경 중 오류: {str(e)}")
            return jsonify({'success': False, 'error': str(e)})
    
    return jsonify({'success': False, 'error': '관리자 계정은 변경할 수 없습니다.'})

@app.route('/admin/extend_subscription/<int:user_id>', methods=['POST'])
@admin_required
def extend_subscription(user_id):
    user = User.query.get_or_404(user_id)
    if not user.is_admin:
        try:
            data = request.get_json()
            print(f"[DEBUG] extend_subscription 요청 데이터: {data}")
            
            is_premium = data.get('is_premium', False)
            period = data.get('period')
            
            # 유료 회원으로 설정
            if is_premium:
                print(f"[DEBUG] 유료 회원으로 설정: {user.username}")
                user.is_premium = True
                user.subscription_expiry = datetime.utcnow() + timedelta(days=36500)  # 약 100년
                flash(f'{user.username} 님을 유료 회원으로 변경했습니다.', 'success')
            # 무료 체험으로 설정 또는 연장
            elif period:
                print(f"[DEBUG] 무료 체험 설정/연장: {user.username}, 기간: {period}")
                # 유료 회원이었다면 무료 체험으로 전환
                if user.is_premium:
                    user.is_premium = False
                    print(f"[DEBUG] 유료 회원에서 무료 체험으로 전환: {user.username}")
                    flash(f'{user.username} 님을 무료 체험 회원으로 변경했습니다.', 'success')
                
                current_time = datetime.utcnow()
                
                if period == '1일로 초기화':
                    # 정확히 24시간으로 초기화
                    user.subscription_expiry = current_time + timedelta(hours=24)
                    flash(f'{user.username} 님의 구독을 24시간으로 초기화했습니다.', 'success')
                elif period == '1일':
                    if user.subscription_expiry and user.subscription_expiry > current_time:
                        # 기존 만료일에 24시간 추가
                        user.subscription_expiry += timedelta(hours=24)
                        flash(f'{user.username} 님의 구독을 24시간 연장했습니다.', 'success')
                    else:
                        # 현재 시간부터 24시간 설정
                        user.subscription_expiry = current_time + timedelta(hours=24)
                        flash(f'{user.username} 님의 구독을 24시간으로 설정했습니다.', 'success')
                elif period == '7일':
                    if user.subscription_expiry and user.subscription_expiry > current_time:
                        # 기존 만료일에 7일(168시간) 추가
                        user.subscription_expiry += timedelta(hours=168)
                        flash(f'{user.username} 님의 구독을 7일(168시간) 연장했습니다.', 'success')
                    else:
                        # 현재 시간부터 7일(168시간) 설정
                        user.subscription_expiry = current_time + timedelta(hours=168)
                        flash(f'{user.username} 님의 구독을 7일(168시간)으로 설정했습니다.', 'success')
                elif period == '30일':
                    if user.subscription_expiry and user.subscription_expiry > current_time:
                        # 기존 만료일에 30일(720시간) 추가
                        user.subscription_expiry += timedelta(hours=720)
                        flash(f'{user.username} 님의 구독을 30일(720시간) 연장했습니다.', 'success')
                    else:
                        # 현재 시간부터 30일(720시간) 설정
                        user.subscription_expiry = current_time + timedelta(hours=720)
                        flash(f'{user.username} 님의 구독을 30일(720시간)으로 설정했습니다.', 'success')
            else:
                print(f"[ERROR] 잘못된 요청 - is_premium 또는 period가 없음: {data}")
                return jsonify({'success': False, 'error': '유효하지 않은 요청입니다. 구독 유형이나 기간을 지정해주세요.'})
            
            db.session.commit()
            print(f"[SUCCESS] 사용자 구독 정보 업데이트 완료: {user.username}, premium={user.is_premium}, expiry={user.subscription_expiry}")
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] 구독 연장 중 오류: {str(e)}")
            return jsonify({'success': False, 'error': f'오류가 발생했습니다: {str(e)}'})
    
    return jsonify({'success': False, 'error': '관리자 계정은 변경할 수 없습니다.'})

@app.route('/admin/toggle_account_status/<int:user_id>')
@admin_required
def toggle_account_status(user_id):
    user = User.query.get_or_404(user_id)
    user.account_status = '정지' if user.account_status == '활성' else '활성'
    db.session.commit()
    return jsonify({'success': True, 'new_status': user.account_status})

@app.route('/admin/get_user/<int:user_id>')
@admin_required
def get_user(user_id):
    user = User.query.get_or_404(user_id)
    return jsonify({
        'success': True,
        'isPremium': user.is_premium,
        'username': user.username,
        'subscriptionExpiry': user.subscription_expiry.strftime('%Y-%m-%d %H:%M:%S') if user.subscription_expiry else None,
        'accountStatus': user.account_status
    })

def create_admin():
    """Create an admin user if it doesn't exist."""
    try:
        admin = User.query.filter_by(email='admin@example.com').first()
        if not admin:
            admin = User(
                username='Administrator',
                email='admin@example.com',
                password='deok3094',
                is_admin=True
            )
            db.session.add(admin)
            db.session.commit()
            print("관리자 계정이 생성되었습니다.")
        else:
            print("관리자 계정이 이미 존재합니다.")
    except Exception as e:
        print(f"관리자 계정 생성 중 오류 발생: {str(e)}")
        db.session.rollback()

@app.route('/account')
@login_required
def account():
    return render_template('account.html')

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not current_user.check_password(current_password):
        flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
        return redirect(url_for('account'))
        
    if new_password != confirm_password:
        flash('새 비밀번호가 일치하지 않습니다.', 'danger')
        return redirect(url_for('account'))
        
    if len(new_password) < 8:
        flash('비밀번호는 8자 이상이어야 합니다.', 'danger')
        return redirect(url_for('account'))
        
    current_user.password = generate_password_hash(new_password)
    db.session.commit()
    flash('비밀번호가 성공적으로 변경되었습니다.', 'success')
    return redirect(url_for('account'))

def get_complex_info(complex_no):
    """아파트 단지 상세 정보 조회"""
    try:
        cookies = {
            'NNB': 'FGYNFS4Y6M6WO',
            'NFS': '2',
            'ASID': 'afd10077000001934e8033f50000004e',
            'ba.uuid': 'a5e52e8f-1775-4eea-9b42-30223205f9df',
            'tooltipDisplayed': 'true',
            'nstore_session': 'zmRE1M3UHwL1GmMzBg0gfcKH',
            '_fwb': '242x1Ggncj6Dnv0G6JF6g8h.1738045585397',
            'landHomeFlashUseYn': 'N',
            'REALESTATE': 'Thu Apr 03 2025 20:14:11 GMT+0900 (Korean Standard Time)',
            'NACT': '1',
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
            'Referer': f'https://new.land.naver.com/complexes/{complex_no}',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3MzgwNDcxNjMsImV4cCI6MTczODA1Nzk2M30.Heq-J33LY9pJDnYOqmRhTTrSPqCpChtWxka_XUphnd4',
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        }
        
        session = requests.Session()
        session.headers.update(headers)
        session.cookies.update(cookies)
        
        init_url = f'https://new.land.naver.com/complexes/{complex_no}'
        init_response = session.get(init_url, verify=False, timeout=10)
        print(f"[DEBUG] Initial request status: {init_response.status_code}")
        if init_response.status_code != 200:
            print(f"[ERROR] Initial request failed: {init_response.text[:500]}")
            return None
            
        url = f'https://new.land.naver.com/api/complexes/overview/{complex_no}'
        response = session.get(url, verify=False, timeout=10)
        print(f"[DEBUG] Complex info API status: {response.status_code}")
        if response.status_code != 200:
            print(f"[ERROR] Complex info request failed: {response.text[:500]}")
            return None
            
        data = response.json()
        print(f"[DEBUG] Complex info response: {data}")
        return data
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error in get_complex_info: {str(e)}, response: {response.text[:500] if 'response' in locals() else 'N/A'}")
        return None
    except Exception as e:
        print(f"[ERROR] Error in get_complex_info: {str(e)}")
        return None

def get_article_list(complex_no, trade_type='', page=1):
    """매물 목록 조회"""
    try:
        cookies = {
            'NNB': 'FGYNFS4Y6M6WO',
            'NFS': '2',
            'ASID': 'afd10077000001934e8033f50000004e',
            'ba.uuid': 'a5e52e8f-1775-4eea-9b42-30223205f9df',
            'tooltipDisplayed': 'true',
            'nstore_session': 'zmRE1M3UHwL1GmMzBg0gfcKH',
            '_fwb': '242x1Ggncj6Dnv0G6JF6g8h.1738045585397',
            'landHomeFlashUseYn': 'N',
            'REALESTATE': 'Thu Apr 03 2025 20:14:11 GMT+0900 (Korean Standard Time)',
            'NACT': '1',
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
            'Referer': f'https://new.land.naver.com/complexes/{complex_no}',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3MzgwNDcxNjMsImV4cCI6MTczODA1Nzk2M30.Heq-J33LY9pJDnYOqmRhTTrSPqCpChtWxka_XUphnd4',
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        }
        
        session = requests.Session()
        session.headers.update(headers)
        session.cookies.update(cookies)
        
        init_url = f'https://new.land.naver.com/complexes/{complex_no}'
        init_response = session.get(init_url, verify=False, timeout=10)
        print(f"[DEBUG] Initial request status: {init_response.status_code}")
        if init_response.status_code != 200:
            print(f"[ERROR] Initial request failed: {init_response.text[:500]}")
            return None
            
        all_articles = []
        current_page = 1
        
        while True:
            url = f'https://new.land.naver.com/api/articles/complex/{complex_no}'
            params = {
                'realEstateType': 'APT:ABYG:JGC:PRE',
                'tradeType': trade_type,
                'tag': '::::::::',
                'rentPriceMin': '0',
                'rentPriceMax': '900000000',
                'priceMin': '0',
                'priceMax': '900000000',
                'areaMin': '0',
                'areaMax': '900000000',
                'showArticle': 'false',
                'sameAddressGroup': 'false',
                'priceType': 'RETAIL',
                'page': str(current_page),
                'complexNo': str(complex_no),
                'type': 'list',
                'order': 'rank'
            }
            
            response = session.get(url, params=params, verify=False, timeout=10)
            print(f"[DEBUG] Article list API status (page {current_page}): {response.status_code}")
            if response.status_code != 200:
                print(f"[ERROR] Article list request failed: {response.text[:500]}")
                return None
                
            try:
                data = response.json()
            except json.JSONDecodeError as e:
                print(f"[ERROR] JSON decode error in get_article_list: {str(e)}, response: {response.text[:500]}")
                return None
                
            articles = data.get('articleList', [])
            if not articles:
                print(f"[DEBUG] No articles on page {current_page}")
                break
                
            all_articles.extend(articles)
            
            if not data.get('isMoreData', False):
                print(f"[DEBUG] No more data after page {current_page}")
                break
                
            current_page += 1
            time.sleep(1)
            
        return {'articleList': all_articles}
        
    except Exception as e:
        print(f"[ERROR] Error in get_article_list: {str(e)}")
        return None

def init_db():
    """안전하게 데이터베이스 초기화"""
    with app.app_context():
        # 개발 환경에서만 기존 데이터베이스를 초기화
        if os.environ.get('FLASK_ENV') != 'production':
            db.drop_all()
            print("개발 환경: 데이터베이스 초기화됨")
        
        # 테이블 생성 (존재하지 않는 경우에만)
        db.create_all()
        
        # 관리자 계정 생성
        create_admin()
        print("데이터베이스 설정 완료")

if __name__ == '__main__':
    # 데이터베이스 초기화
    init_db()
    
    # 로컬 개발 환경
    host = '0.0.0.0'  # 모든 인터페이스에서 접속 허용
    port = 5000
    
    print(f"서버가 시작되었습니다. 로컬 접속 주소: http://127.0.0.1:{port}")
    print(f"네트워크 접속 주소: http://{host}:{port}")
    print(f"내부 IP: http://192.168.68.101:{port}")
    
    app.run(host=host, port=port, debug=True)
